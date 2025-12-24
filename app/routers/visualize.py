# backend/services/data_manager.py

from backend.database import SessionLocal
from backend.models.report import Report
from backend.models.inference import Inference
from backend.models.user_settings import UserSettings
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi import APIRouter, Request, Form, UploadFile, File, BackgroundTasks
from fastapi.templating import Jinja2Templates
from sqlalchemy import func, text
from sqlalchemy.exc import OperationalError
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


def get_db_session():
    """Get a database session with automatic retry on connection failure"""
    max_retries = 3
    retry_delay = 0.5
    
    for attempt in range(max_retries):
        try:
            db = SessionLocal()
            # Test the connection
            db.execute(text("SELECT 1"))
            return db
        except OperationalError as e:
            if attempt < max_retries - 1:
                import time
                time.sleep(retry_delay)
            else:
                raise


@router.get("/visualize", response_class=HTMLResponse)
def visualize_reports(request: Request, search: str = None, date: str = None, report: int = None):
    # Check if user is logged in
    user_id = request.session.get("user_id")
    if not user_id:
        return RedirectResponse("/login", status_code=303)
    
    db = None
    try:
        db = get_db_session()
        
        # Get user settings with retry
        user_settings = db.query(UserSettings).filter(UserSettings.user_id == user_id).first()
        if not user_settings:
            # Create default settings if not exists
            user_settings = UserSettings(user_id=user_id)
            db.add(user_settings)
            db.commit()
        
        # Start with base query - FILTER BY USER_ID
        query = db.query(Report).filter(Report.user_id == user_id).order_by(Report.createdAt.desc())
        
        # Apply date filter if provided
        if date:
            try:
                filter_date = datetime.strptime(date, "%Y-%m-%d").date()
                query = query.filter(Report.createdAt == filter_date)
            except ValueError:
                pass
        
        # Get filtered reports
        reports = query.all()
        
        # Get unique dates for date filter (sorted descending) - FILTERED BY USER_ID
        unique_dates = db.query(Report.createdAt).filter(Report.user_id == user_id).distinct().filter(Report.createdAt.isnot(None)).order_by(Report.createdAt.desc()).all()
        unique_dates = [str(d[0]) if d[0] else None for d in unique_dates if d[0]]
        
        # Get auto-select report if report ID is provided
        selected_report_id = report
        selected_report_data = None
        if selected_report_id:
            selected_report = db.query(Report).filter(
                Report.id == selected_report_id,
                Report.user_id == user_id  # ENSURE OWNER MATCHES
            ).first()
            if selected_report:
                inferences = db.query(Inference).filter(Inference.report_id == selected_report_id).order_by(Inference.id.desc()).all()
                
                # Calculate levels based on user settings and number of inferences
                inferences_with_levels = []
                images_per_row = user_settings.images_per_row
                level_prefix = user_settings.level_prefix
                
                for idx, inf in enumerate(inferences):
                    level_number = (idx // images_per_row) + 1
                    position_in_level = (idx % images_per_row) + 1
                    level_name = f"{level_prefix}{level_number}-{position_in_level}"
                    
                    inferences_with_levels.append({
                        "id": inf.id,
                        "unique_id": inf.unique_id,
                        "vin_no": inf.vin_no,
                        "quantity": inf.quantity,
                        "image_name": inf.image_name,
                        "s3_obj_url": inf.s3_obj_url,
                        "exclusion": inf.exclusion,
                        "is_non_confirmity": inf.is_non_confirmity,
                        "createdAt": str(inf.createdAt),
                        "level_name": level_name,
                        "level_number": level_number,
                        "position_in_level": position_in_level
                    })
                
                selected_report_data = {
                    "id": selected_report.id,
                    "report_name": selected_report.report_name,
                    "createdAt": str(selected_report.createdAt),
                    "inferences": inferences_with_levels
                }
        
        return templates.TemplateResponse("visualize.html", {
            "request": request,
            "reports": reports,
            "unique_dates": unique_dates,
            "selected_date": date,
            "search_query": search,
            "selected_report": selected_report_data,
            "selected_report_id": selected_report_id,
            "user_settings": {
                "images_per_row": user_settings.images_per_row,
                "level_prefix": user_settings.level_prefix,
                "image_size": user_settings.image_size,
                "show_image_info": user_settings.show_image_info,
                "show_level_info": user_settings.show_level_info
            }
        })
    except OperationalError as e:
        logger.error(f"Database connection error in visualize_reports: {str(e)}")
        return templates.TemplateResponse("visualize.html", {
            "request": request,
            "reports": [],
            "unique_dates": [],
            "error": "Database connection error. Please refresh the page."
        })
    except Exception as e:
        logger.error(f"Error in visualize_reports: {str(e)}")
        return templates.TemplateResponse("visualize.html", {
            "request": request,
            "reports": [],
            "unique_dates": [],
            "error": f"Error loading reports. Please try again."
        })
    finally:
        if db:
            db.close()

@router.get("/api/report/{report_id}/details")
def get_report_details_api(request: Request, report_id: int):
    # Check if user is logged in and owns this report
    user_id = request.session.get("user_id")
    if not user_id:
        return {"error": "Unauthorized"}
    
    db = None
    try:
        db = get_db_session()
        
        # Get user settings
        user_settings = db.query(UserSettings).filter(UserSettings.user_id == user_id).first()
        if not user_settings:
            user_settings = UserSettings(user_id=user_id)
            db.add(user_settings)
            db.commit()
        
        report = db.query(Report).filter(
            Report.id == report_id,
            Report.user_id == user_id  # ENSURE OWNER MATCHES
        ).first()
        if not report:
            return {"error": "Report not found"}
        
        inferences = db.query(Inference).filter(Inference.report_id == report_id).all()
        
        # Calculate levels based on user settings
        images_per_row = user_settings.images_per_row
        level_prefix = user_settings.level_prefix
        
        inferences_with_levels = []
        for idx, inf in enumerate(inferences):
            level_number = (idx // images_per_row) + 1
            position_in_level = (idx % images_per_row) + 1
            level_name = f"{level_prefix}{level_number}-{position_in_level}"
            
            inferences_with_levels.append({
                "id": inf.id,
                "unique_id": inf.unique_id,
                "vin_no": inf.vin_no,
                "quantity": inf.quantity,
                "image_name": inf.image_name,
                "s3_obj_url": inf.s3_obj_url,
                "exclusion": inf.exclusion,
                "is_non_confirmity": inf.is_non_confirmity,
                "createdAt": str(inf.createdAt),
                "level_name": level_name,
                "level_number": level_number,
                "position_in_level": position_in_level
            })
        
        return {
            "id": report.id,
            "report_name": report.report_name,
            "createdAt": str(report.createdAt),
            "inferences": inferences_with_levels,
            "user_settings": {
                "images_per_row": user_settings.images_per_row,
                "level_prefix": user_settings.level_prefix,
                "image_size": user_settings.image_size,
                "show_image_info": user_settings.show_image_info,
                "show_level_info": user_settings.show_level_info
            }
        }
    except OperationalError as e:
        logger.error(f"Database connection error in get_report_details_api: {str(e)}")
        return {"error": "Database connection error. Please refresh the page."}
    except Exception as e:
        logger.error(f"Error in get_report_details_api: {str(e)}")
        return {"error": f"Error loading report details."}
    finally:
        if db:
            db.close()


# --------------------------
#  REPORTS
# --------------------------

def get_reports(date=None):
    db = SessionLocal()
    try:
        if date:
            return db.query(Report).filter(Report.createdAt == date).all()
        return db.query(Report).all()
    finally:
        db.close()


def get_report(report_id: int):
    db = SessionLocal()
    try:
        return db.query(Report).filter(Report.id == report_id).first()
    finally:
        db.close()


def create_report(report_name: str):
    db = SessionLocal()
    try:
        report = Report(report_name=report_name)
        db.add(report)
        db.commit()
        db.refresh(report)
        return report.id
    finally:
        db.close()


def delete_report(report_id: int):
    db = SessionLocal()
    try:
        report = db.query(Report).filter(Report.id == report_id).first()
        if report:
            db.delete(report)
            db.commit()
    finally:
        db.close()


# --------------------------
#  INFERENCE HELPERS
# --------------------------

def upload_result(inference_obj: Inference):
    db = SessionLocal()
    try:
        db.add(inference_obj)
        db.commit()
        db.refresh(inference_obj)
        return inference_obj
    finally:
        db.close()


def get_report_details(report_id: int):
    db = SessionLocal()
    try:
        return db.query(Inference).filter(Inference.report_id == report_id).all()
    finally:
        db.close()


@router.get("/api/report/{report_id}/export/excel")
def export_report_excel(request: Request, report_id: int):
    """Export report data to Excel file with summary and pie chart"""
    # Check if user is logged in and owns this report
    user_id = request.session.get("user_id")
    if not user_id:
        return {"error": "Unauthorized"}
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')  # Use non-GUI backend
    
    db = None
    try:
        db = get_db_session()
        
        report = db.query(Report).filter(
            Report.id == report_id,
            Report.user_id == user_id  # ENSURE OWNER MATCHES
        ).first()
        if not report:
            return {"error": "Report not found"}
        
        inferences = db.query(Inference).filter(
            Inference.report_id == report_id
        ).order_by(Inference.id.desc()).all()
        
        # Calculate summary statistics
        summary = calculate_summary_for_export(inferences)
        
        # Get user settings for location calculation
        user_settings = db.query(UserSettings).filter(UserSettings.user_id == user_id).first()
        if not user_settings:
            user_settings = UserSettings(user_id=user_id)
        
        # Calculate locations for each inference
        images_per_row = user_settings.images_per_row
        level_prefix = user_settings.level_prefix
        inferences_with_locations = []
        
        for idx, inf in enumerate(inferences):
            level_number = (idx // images_per_row) + 1
            position_in_level = (idx % images_per_row) + 1
            level_name = f"{level_prefix}{level_number}-{position_in_level}"
            inferences_with_locations.append({
                'inference': inf,
                'level_name': level_name,
                'level_number': level_number,
                'position': position_in_level
            })

        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Data"
        
        # Define styles
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        summary_fill = PatternFill(start_color="e8eef7", end_color="e8eef7", fill_type="solid")
        summary_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ====== SECTION 1: Report Header ======
        ws['A1'] = f"Report: {report.report_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Created: {report.createdAt}"
        ws['A2'].font = Font(size=11)
        ws['A3'] = f"Total Items: {len(inferences)}"
        ws['A3'].font = Font(size=11)
        
        # ====== SECTION 2: Summary Statistics ======
        current_row = 5
        ws[f'A{current_row}'] = "📊 SUMMARY STATISTICS"
        ws[f'A{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{current_row}'].fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        ws.merge_cells(f'A{current_row}:B{current_row}')
        
        current_row += 1
        
        # Summary items
        summary_items = [
            ("✓ Filled", summary.get("filled", 0)),
            ("⚪ Empty Skid", summary.get("empty", 0)),
            ("❌ Sticker Not Found", summary.get("stickerNotFound", 0)),
            ("⚠️ Multiple Stickers", summary.get("multipleStickers", 0)),
            ("📝 Other", summary.get("other", 0))
        ]
        
        for label, count in summary_items:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].fill = summary_fill
            ws[f'A{current_row}'].font = summary_font
            ws[f'B{current_row}'] = count
            ws[f'B{current_row}'].fill = summary_fill
            ws[f'B{current_row}'].font = summary_font
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center")
            current_row += 1
        
        # ====== SECTION 3: Pie Chart ======
        # Create pie chart image
        chart_img_path = None
        try:
            fig, ax = plt.subplots(figsize=(8, 6))
            labels = [item[0] for item in summary_items]
            sizes = [item[1] for item in summary_items]
            colors = ['#28a745', '#ffc107', '#e83e8c', '#fd7e14', '#6c757d']
            
            # Filter out zero values
            non_zero_labels = [l for l, s in zip(labels, sizes) if s > 0]
            non_zero_sizes = [s for s in sizes if s > 0]
            non_zero_colors = [c for c, s in zip(colors, sizes) if s > 0]
            
            if non_zero_sizes:
                ax.pie(non_zero_sizes, labels=non_zero_labels, colors=non_zero_colors, 
                       autopct='%1.1f%%', startangle=90, textprops={'fontsize': 10})
                ax.set_title(f"Item Distribution - {report.report_name}", fontweight='bold', fontsize=12)
                
                # Save chart to bytes
                chart_img_path = BytesIO()
                plt.savefig(chart_img_path, format='png', bbox_inches='tight', dpi=100)
                chart_img_path.seek(0)
                plt.close(fig)
                
                # Insert chart into Excel
                current_row += 2
                ws[f'A{current_row}'] = "📈 CHART"
                ws[f'A{current_row}'].font = Font(bold=True, size=11)
                
                current_row += 1
                chart_xl_image = XLImage(chart_img_path)
                chart_xl_image.width = 400
                chart_xl_image.height = 300
                ws.add_image(chart_xl_image, f'A{current_row}')
                
                current_row += 16  # Approximate rows taken by image
            else:
                plt.close(fig)
        except Exception as e:
            logger.warning(f"Could not generate pie chart for Excel export: {str(e)}")
        
        # ====== SECTION 4: Data Table ======
        current_row += 2
        ws[f'A{current_row}'] = "📋 DETAILED DATA"
        ws[f'A{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{current_row}'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        ws.merge_cells(f'A{current_row}:J{current_row}')

        
        current_row += 1
        
        # Define headers
        headers = ["Item #", "Location", "Unique ID", "VIN Number", "Quantity", "Image Name", "Exclusion", "Created Date", "Download Image", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Add data rows
        for idx, item in enumerate(inferences_with_locations, 1):
            row = current_row + idx
            inference = item['inference']
            status = "Non-Conformity" if inference.is_non_confirmity else "Confirmed"
            data = [
                idx,
                item['level_name'],  # Location (e.g., L1-1)
                inference.unique_id or "",
                inference.vin_no or "",
                inference.quantity or 0,
                inference.image_name or "",
                inference.exclusion or "",
                str(inference.createdAt) if inference.createdAt else "",
                inference.s3_obj_url or "",
                status
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # Make S3 URL a hyperlink
                if col == 9 and inference.s3_obj_url:
                    cell.hyperlink = inference.s3_obj_url
                    cell.value = "Download"
                    cell.font = Font(color="0563C1", underline="single")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12  # Location (L1-1)
        ws.column_dimensions['C'].width = 20  # Unique ID
        ws.column_dimensions['D'].width = 20  # VIN
        ws.column_dimensions['E'].width = 10  # Quantity
        ws.column_dimensions['F'].width = 25  # Image Name
        ws.column_dimensions['G'].width = 18  # Exclusion
        ws.column_dimensions['H'].width = 18  # Created Date
        ws.column_dimensions['I'].width = 15  # Download
        ws.column_dimensions['J'].width = 15  # Status

        ws.column_dimensions['I'].width = 15
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return StreamingResponse(
            iter([excel_file.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Report_{report.id}_{report.report_name.replace(' ', '_')}.xlsx"}
        )
    except OperationalError as e:
        logger.error(f"Database connection error in export_report_excel: {str(e)}")
        return {"error": "Database connection error. Please try again."}
    except Exception as e:
        logger.error(f"Error in export_report_excel: {str(e)}")
        return {"error": f"Error exporting report."}
    finally:
        if db:
            db.close()


def calculate_summary_for_export(inferences):
    """Calculate summary statistics from inferences"""
    summary = {
        "filled": 0,
        "empty": 0,
        "stickerNotFound": 0,
        "multipleStickers": 0,
        "other": 0
    }
    
    for inf in inferences:
        if inf.exclusion:
            exclusion_lower = inf.exclusion.lower()
            if 'empty' in exclusion_lower:
                summary["empty"] += 1
            elif 'sticker' in exclusion_lower and 'not' in exclusion_lower:
                summary["stickerNotFound"] += 1
            elif 'multiple' in exclusion_lower:
                summary["multipleStickers"] += 1
            elif 'filled' in exclusion_lower:
                summary["filled"] += 1
            else:
                summary["other"] += 1
        else:
            summary["filled"] += 1
    
    return summary

